from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Reuse the SAME helpers the Member Schedule PDF uses, so the workbook values, PASS/FAIL
# status, off-scale ("> 10.00") handling, "N/A" suppression and summary statistics match
# the PDF exactly. No recalculation is performed — saved member data is the source of truth.
from reports.generators.member_schedule_report import (
    _BEAM_LABEL,
    _COLUMN_LABEL,
    _OFF_SCALE_UTIL,
    _format_util_display,
    _is_grossly_invalid,
    _is_passing,
    _status_label,
    _text,
    _to_float,
)


@dataclass(frozen=True)
class GeneratedWorkbook:
    filename: str
    content: bytes
    media_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _num_or_text(value: Any) -> Any:
    """Numeric cell when the saved value parses to a number (enables sorting/filtering),
    otherwise the original text. Numeric value is identical to the saved display value."""
    number = _to_float(value)
    return number if number is not None else _text(value)


def _excel_util(member: dict) -> Any:
    """Control utilization matching the PDF: numeric when on-scale, '> 10.00' when off-scale."""
    util = _to_float(member.get("controlValue"))
    if util is None:
        return _text(member.get("controlValue"))
    if util > _OFF_SCALE_UTIL:
        return _format_util_display(member.get("controlValue"))  # "> 10.00"
    return util


def _slof_or_na(field: str) -> Callable[[dict], Any]:
    def getter(member: dict) -> Any:
        if _is_grossly_invalid(member):
            return "N/A"
        return _num_or_text(member.get(field))
    return getter


# Column spec: (header, value_fn, number_format_or_None). Adding a new schedule type is a
# matter of appending a SCHEDULE_SHEETS entry — the writer below is fully generic.
_BEAM_COLUMNS = [
    ("Member", lambda m: _text(m.get("name")), None),
    ("Length (m)", lambda m: _num_or_text(m.get("length")), "0.000"),
    ("Selected Section", lambda m: _text(m.get("section")), None),
    ("δmax (mm)", lambda m: _text(m.get("deflectionValue")), None),
    ("Control Utilization", _excel_util, "0.000"),
    ("Left Slof Width (cm)", lambda m: _num_or_text(m.get("leftSlofWidth")), "0.00"),
    ("Left Slof Length (cm)", _slof_or_na("leftSlofLength"), "0.00"),
    ("Right Slof Width (cm)", lambda m: _num_or_text(m.get("rightSlofWidth")), "0.00"),
    ("Right Slof Length (cm)", _slof_or_na("rightSlofLength"), "0.00"),
    ("Middle Reinforcement (cm²)", _slof_or_na("reinforcementMid"), "0.000"),
    ("End Reinforcement (cm²)", _slof_or_na("reinforcementHead"), "0.000"),
    ("Status", _status_label, None),
]

_COLUMN_COLUMNS = [
    ("Member", lambda m: _text(m.get("name")), None),
    ("Length (m)", lambda m: _num_or_text(m.get("length")), "0.000"),
    ("Selected Section", lambda m: _text(m.get("section")), None),
    ("Control Utilization", _excel_util, "0.000"),
    ("Governing Axis", lambda m: _text(m.get("governingAxis")), None),
    ("Status", _status_label, None),
]

SCHEDULE_SHEETS = [
    {"title": "Beam Schedule", "label": _BEAM_LABEL, "columns": _BEAM_COLUMNS},
    {"title": "Column Schedule", "label": _COLUMN_LABEL, "columns": _COLUMN_COLUMNS},
]

_HEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
_HEADER_FONT = Font(bold=True)


def _write_schedule_sheet(ws, columns: list, members: list[dict]) -> None:
    headers = [c[0] for c in columns]
    ws.append(headers)
    for member in members:
        ws.append([fn(member) for (_h, fn, _fmt) in columns])

    # Header styling.
    for col_index, _header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Per-cell number formats and auto-sized columns.
    for col_index, (_h, _fn, number_format) in enumerate(columns, start=1):
        letter = get_column_letter(col_index)
        max_len = len(str(headers[col_index - 1]))
        for row_index in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_index, column=col_index)
            if number_format and isinstance(cell.value, (int, float)):
                cell.number_format = number_format
            max_len = max(max_len, len(str(cell.value if cell.value is not None else "")))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)

    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"


def _write_summary_sheet(ws, *, project_name: str, created: datetime, beams: list[dict], columns: list[dict]) -> None:
    relevant = beams + columns
    # Identical arithmetic to the Member Schedule PDF summary.
    utils = [u for u in (_to_float(m.get("controlValue")) for m in relevant) if u is not None]
    on_scale = [u for u in utils if u <= _OFF_SCALE_UTIL]
    flags = [_is_passing(m) for m in relevant]
    passing = sum(1 for f in flags if f is True)
    failing = sum(1 for f in flags if f is False)
    max_util = _format_util_display(max(utils)) if utils else "-"
    avg_util = f"{(sum(on_scale) / len(on_scale)):.3f}" if on_scale else "-"

    rows = [
        ("Project Name", project_name or "Not specified"),
        ("Export Date", created.strftime("%Y-%m-%d %H:%M")),
        ("Total Beams", len(beams)),
        ("Total Columns", len(columns)),
        ("Total Members", len(beams) + len(columns)),
        ("Passing Members", passing),
        ("Failing Members", failing),
        ("Maximum Utilization", max_util),
        ("Average Utilization", avg_util),
    ]
    ws.append(["Summary", ""])
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    for label, value in rows:
        ws.append([label, value])
    for row_index in range(2, ws.max_row + 1):
        ws.cell(row=row_index, column=1).font = _HEADER_FONT
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.freeze_panes = "A2"


def _safe_filename(project_name: str) -> str:
    name = (project_name or "").strip()
    if not name:
        return "Member_Schedule.xlsx"
    cleaned = "".join(ch if (ch.isalnum() or ch in " _-") else "_" for ch in name).strip()
    cleaned = cleaned.replace(" ", "_") or "Member_Schedule"
    return f"{cleaned}.xlsx"


def generate_member_schedule_excel(
    *,
    request_data: dict[str, Any],
    created_at: datetime | None = None,
) -> GeneratedWorkbook:
    created = created_at or datetime.now()
    project_name = _text(request_data.get("project_name"), "")
    if project_name == "-":
        project_name = ""
    members = request_data.get("members") or []

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    beams = [m for m in members if str(m.get("modeLabel")) == _BEAM_LABEL]
    columns = [m for m in members if str(m.get("modeLabel")) == _COLUMN_LABEL]
    by_label = {_BEAM_LABEL: beams, _COLUMN_LABEL: columns}

    for spec in SCHEDULE_SHEETS:
        ws = wb.create_sheet(title=spec["title"])
        sheet_members = by_label.get(spec["label"], [m for m in members if str(m.get("modeLabel")) == spec["label"]])
        _write_schedule_sheet(ws, spec["columns"], sheet_members)

    summary_ws = wb.create_sheet(title="Summary")
    _write_summary_sheet(summary_ws, project_name=project_name, created=created, beams=beams, columns=columns)

    buffer = BytesIO()
    wb.save(buffer)
    content = buffer.getvalue()
    if not content.startswith(b"PK"):  # .xlsx is a zip (OOXML) container
        raise RuntimeError("Generated member schedule workbook is not a valid XLSX stream.")

    return GeneratedWorkbook(filename=_safe_filename(project_name), content=content)
